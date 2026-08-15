import os
import time
import requests
import openpyxl

# 配置
XLSX_PATH = os.path.join(os.path.dirname(__file__), 'mac_actor.xlsx')
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'upload')

def extract_path(url):
    """从 actor_pic URL 中提取 /img/ 后面的路径，去掉查询参数"""
    url = url.split('?')[0]
    marker = '/img/'
    idx = url.find(marker)
    if idx == -1:
        return None
    return url[idx + len(marker):]

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    tasks = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        url = row[0].value
        if not url:
            continue
        path = extract_path(url)
        if path:
            tasks.append((url, path))

    print(f'共 {len(tasks)} 条记录待下载', flush=True)

    success, skip, fail = 0, 0, 0
    for i, (url, path) in enumerate(tasks, 1):
        save_path = os.path.join(UPLOAD_DIR, path)

        # 已存在则跳过
        if os.path.exists(save_path):
            skip += 1
            continue

        # 确保目录存在
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        # 下载图片
        full_url = 'http://' + url.split('?')[0]
        try:
            resp = requests.get(full_url, timeout=10)
            resp.raise_for_status()
            with open(save_path, 'wb') as f:
                f.write(resp.content)
            print(f'  [{i}/{len(tasks)}] 成功: {path}', flush=True)
        except Exception as e:
            fail += 1
            print(f'  [{i}/{len(tasks)}] 失败: {full_url} -> {e}', flush=True)

        time.sleep(1)

    print(f'\n完成! 成功:{success} 跳过:{skip} 失败:{fail}')

if __name__ == '__main__':
    main()
